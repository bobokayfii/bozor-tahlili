"""Kategoriyalarni chiroyli formatlangan Excel (.xlsx) fayliga eksport
qilish — bitta joriy kategoriya (bitta varaq) yoki barcha kategoriyalar
(har biri o'z varag'ida, bitta faylda) sifatida. Ustunlar frontend'dagi
getProductColumns bilan bir xil mantiqqa amal qiladi (mikroqarz —
qisqartirilgan, credit_special_terms — Maxsus shartlari bilan, aks holda
Boshlang'ich badal bilan), shuning uchun yuklab olingan fayl ekrandagi
jadvalga mos keladi."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from db.models import ProductRow
from unavailable_products import UnavailableBank

_HEADER_FILL = PatternFill(start_color="FF00355F", end_color="FF00355F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_DASH = "—"

_TEXT = {
    "uz": {
        "rank": "#",
        "bank": "Bank",
        "product": "Mahsulot",
        "rate": "Stavka",
        "term": "Muddat",
        "down_payment": "Boshlang'ich badal",
        "grace_period": "Imtiyozli davr",
        "amount": "Kredit miqdori",
        "special_terms": "Maxsus shartlari",
        "payment_method": "To'lov usuli",
        "reason": "Sabab",
        "yes": "Bor",
        "no": "Yo'q",
        "month_unit": "oy",
        "amount_unit": "mln so'm",
        "default_payment_method": "Annuitet, Differensial",
    },
    "ru": {
        "rank": "#",
        "bank": "Банк",
        "product": "Продукт",
        "rate": "Ставка",
        "term": "Срок",
        "down_payment": "Первоначальный взнос",
        "grace_period": "Льготный период",
        "amount": "Сумма кредита",
        "special_terms": "Особые условия",
        "payment_method": "Способ оплаты",
        "reason": "Причина",
        "yes": "Есть",
        "no": "Нет",
        "month_unit": "мес.",
        "amount_unit": "млн сум",
        "default_payment_method": "Аннуитет, Дифференцированный",
    },
}

_PAYMENT_METHOD_RU = {
    "Annuitet": "Аннуитет",
    "Differensial": "Дифференцированный",
    "Annuitet, Differensial": "Аннуитет, Дифференцированный",
}


def _rate_text(product: ProductRow) -> str:
    if product.rate_min == product.rate_max:
        return f"{product.rate_min}%"
    return f"{product.rate_min}%–{product.rate_max}%"


def _term_text(product: ProductRow, text: dict[str, str]) -> str:
    unit = text["month_unit"]
    if product.term_min_months == product.term_max_months:
        return f"{product.term_min_months} {unit}"
    return f"{product.term_min_months}–{product.term_max_months} {unit}"


def _amount_text(product: ProductRow, text: dict[str, str]) -> str:
    millions = round(product.amount_max_som / 1_000_000, 1)
    return f"{millions} {text['amount_unit']}"


def _grace_text(grace_period_months: int | None, text: dict[str, str]) -> str:
    if grace_period_months is None:
        return text["no"]
    return text["yes"] if grace_period_months > 0 else text["no"]


def _payment_method_text(payment_method: str | None, lang: str, text: dict[str, str]) -> str:
    if payment_method is None:
        return text["default_payment_method"]
    if lang != "ru":
        return payment_method
    return _PAYMENT_METHOD_RU.get(payment_method, payment_method)


def _write_category_sheet(
    sheet: Worksheet,
    category_key: str,
    products: list[ProductRow],
    unavailable_banks: list[UnavailableBank],
    schema: str,
    lang: str,
) -> None:
    text = _TEXT.get(lang, _TEXT["uz"])
    is_mikroqarz = category_key in ("mikroqarz", "mikroqarz_onlayn")
    is_special_terms = schema == "credit_special_terms" and not is_mikroqarz

    headers = [text["rank"], text["bank"], text["product"], text["rate"], text["term"]]
    if is_mikroqarz:
        headers += [text["amount"], text["payment_method"]]
    elif is_special_terms:
        headers += [text["amount"], text["grace_period"], text["special_terms"], text["payment_method"]]
    else:
        headers += [text["down_payment"], text["grace_period"], text["amount"], text["payment_method"]]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ranked = sorted(products, key=lambda p: p.rate_min)
    for index, product in enumerate(ranked, start=1):
        row = [index, product.bank, product.product_name, _rate_text(product), _term_text(product, text)]
        if is_mikroqarz:
            row += [_amount_text(product, text), _payment_method_text(product.payment_method, lang, text)]
        elif is_special_terms:
            row += [
                _amount_text(product, text),
                _grace_text(product.grace_period_months, text),
                product.special_terms or _DASH,
                _payment_method_text(product.payment_method, lang, text),
            ]
        else:
            row += [
                f"{product.down_payment_pct}%" if product.down_payment_pct is not None else _DASH,
                _grace_text(product.grace_period_months, text),
                _amount_text(product, text),
                _payment_method_text(product.payment_method, lang, text),
            ]
        sheet.append(row)

    for bank_entry in unavailable_banks:
        filler_count = len(headers) - 3
        sheet.append([None, bank_entry.bank, bank_entry.reason] + [None] * filler_count)

    widths = [4, 20, 34, 14, 12, 18, 14, 16, 26]
    for column_index, width in enumerate(widths[: len(headers)], start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"


def _unique_sheet_title(label: str, used_titles: set[str]) -> str:
    # Excel varaq nomi 31 belgidan oshmasligi kerak va \/:*?[] belgilarini
    # qabul qilmaydi — kategoriya nomlari bunday belgilarni ishlatmaydi,
    # shuning uchun faqat uzunlik cheklanadi. Kelajakda qo'shiladigan
    # kategoriya nomi 31 belgigacha qirqilganda boshqasi bilan tasodifan
    # bir xil chiqib qolsa (hozircha bunday holat yo'q), Excel bir xil
    # nomli ikkita varaqni rad etadi — shu sabab raqamli qo'shimcha bilan
    # ajratiladi.
    base = label[:31]
    if base not in used_titles:
        return base
    counter = 2
    while True:
        suffix = f" ({counter})"
        candidate = label[: 31 - len(suffix)] + suffix
        if candidate not in used_titles:
            return candidate
        counter += 1


def build_category_workbook(
    category_key: str,
    sheet_title: str,
    products: list[ProductRow],
    unavailable_banks: list[UnavailableBank],
    schema: str,
    lang: str = "uz",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_title[:31]
    _write_category_sheet(sheet, category_key, products, unavailable_banks, schema, lang)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_all_categories_workbook(
    categories: list[tuple[str, str, str]],
    products_by_category: dict[str, list[ProductRow]],
    unavailable_by_category: dict[str, list[UnavailableBank]],
    lang: str = "uz",
) -> bytes:
    """`categories` — [(key, label, schema), ...] categories.py'dagi tartibda.
    Har bir kategoriya bitta varaqda, kategoriya nomi bilan nomlangan."""
    workbook = Workbook()
    default_sheet = workbook.active
    assert default_sheet is not None
    workbook.remove(default_sheet)

    used_titles: set[str] = set()
    for key, label, schema in categories:
        title = _unique_sheet_title(label, used_titles)
        used_titles.add(title)
        sheet = workbook.create_sheet(title=title)
        _write_category_sheet(
            sheet,
            key,
            products_by_category.get(key, []),
            unavailable_by_category.get(key, []),
            schema,
            lang,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
