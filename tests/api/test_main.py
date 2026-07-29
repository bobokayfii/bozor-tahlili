from datetime import datetime, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import api.main as api_main
from db.database import get_engine, get_session_factory, init_db
from db.models import ProductRow


@pytest.fixture
def client(tmp_path, monkeypatch):
    engine = get_engine(tmp_path / "api_test.db")
    init_db(engine)
    session_factory = get_session_factory(engine)

    with session_factory() as session:
        session.add(ProductRow(
            bank="SQB", category="mikroqarz", product_name="SQB Mikroqarz",
            rate_min=28.0, rate_max=31.0, term_min_months=3, term_max_months=36,
            amount_max_som=100_000_000, requires_collateral=False,
            down_payment_pct=None, source_url="https://sqb.uz",
            scraped_at=datetime.now(timezone.utc),
        ))
        session.commit()

    monkeypatch.setattr(api_main, "SessionLocal", session_factory)
    return TestClient(api_main.app)


def test_list_products_returns_seeded_row(client):
    response = client.get("/products", params={"category": "mikroqarz"})
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["bank"] == "SQB"


def test_list_products_returns_only_latest_scrape_per_bank_category(client):
    with api_main.SessionLocal() as session:
        session.add(ProductRow(
            bank="HamkorBank", category="avtokredit", product_name="Hamkor Avtokredit (old)",
            rate_min=10.0, rate_max=15.0, term_min_months=12, term_max_months=36,
            amount_max_som=100_000_000, requires_collateral=False,
            down_payment_pct=None, source_url="https://hamkor.uz",
            scraped_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
        ))
        session.add(ProductRow(
            bank="HamkorBank", category="avtokredit", product_name="Hamkor Avtokredit (new)",
            rate_min=20.0, rate_max=25.0, term_min_months=12, term_max_months=36,
            amount_max_som=100_000_000, requires_collateral=False,
            down_payment_pct=None, source_url="https://hamkor.uz",
            scraped_at=datetime(2026, 6, 1, tzinfo=timezone.utc),
        ))
        session.commit()

    response = client.get("/products", params={"category": "avtokredit"})
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["rate_min"] == 20.0
    assert data[0]["product_name"] == "Hamkor Avtokredit (new)"


def test_recommend_uses_only_latest_scrape_per_bank_category(client):
    with api_main.SessionLocal() as session:
        session.add(ProductRow(
            bank="HamkorBank", category="avtokredit", product_name="Hamkor Avtokredit (old)",
            rate_min=10.0, rate_max=15.0, term_min_months=12, term_max_months=36,
            amount_max_som=100_000_000, requires_collateral=False,
            down_payment_pct=None, source_url="https://hamkor.uz",
            scraped_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
        ))
        session.add(ProductRow(
            bank="HamkorBank", category="avtokredit", product_name="Hamkor Avtokredit (new)",
            rate_min=20.0, rate_max=25.0, term_min_months=12, term_max_months=36,
            amount_max_som=100_000_000, requires_collateral=False,
            down_payment_pct=None, source_url="https://hamkor.uz",
            scraped_at=datetime(2026, 6, 1, tzinfo=timezone.utc),
        ))
        session.commit()

    response = client.post("/recommend", json={
        "category": "avtokredit",
        "amount_som": 50_000_000,
        "term_months": 12,
        "collateral_ok": True,
    })
    assert response.status_code == 200
    data = response.json()
    hamkor_matches = [r for r in data["recommendations"] if r["bank"] == "HamkorBank"]
    assert len(hamkor_matches) == 1
    assert hamkor_matches[0]["product_name"] == "Hamkor Avtokredit (new)"


def test_recommend_returns_ranked_list_and_explanation(client, monkeypatch):
    monkeypatch.setattr(
        api_main, "explain_recommendation", lambda criteria, ranked: "test tushuntirish"
    )
    response = client.post("/recommend", json={
        "category": "mikroqarz",
        "amount_som": 50_000_000,
        "term_months": 12,
        "collateral_ok": True,
    })
    assert response.status_code == 200
    data = response.json()
    assert data["explanation"] == "test tushuntirish"
    assert data["recommendations"][0]["bank"] == "SQB"


def test_explain_product_calls_explain_featured_product_with_the_given_product_only(client, monkeypatch):
    """/explain-product /recommend'ning ballash/saralashiga umuman
    tayanmaydi — u faqat so'rovda kelgan (frontend "featured" deb
    tanlagan) bank/mahsulot haqida yozadi, boshqa hech qanday ranking
    ishlatilmaydi."""
    captured = {}

    def fake_explain(category, product, other_bank_count, language="uz"):
        captured["category"] = category
        captured["product"] = product
        captured["other_bank_count"] = other_bank_count
        captured["language"] = language
        return "test tushuntirish"

    monkeypatch.setattr(api_main, "explain_featured_product", fake_explain)

    response = client.post("/explain-product", json={
        "category": "mikroqarz",
        "bank": "HamkorBank",
        "product_name": "Hamkor Mikroqarz",
        "rate_min": 10.0,
        "rate_max": 15.0,
        "term_min_months": 12,
        "term_max_months": 36,
        "amount_max_som": 100_000_000,
        "requires_collateral": False,
        "down_payment_pct": None,
    })

    assert response.status_code == 200
    assert response.json() == {"explanation": "test tushuntirish"}
    assert captured["category"] == "mikroqarz"
    assert captured["product"].bank == "HamkorBank"
    assert captured["product"].product_name == "Hamkor Mikroqarz"
    # Fixture'da faqat SQB Mikroqarz bor — so'ralgan bank (HamkorBank) undan
    # farqli, shuning uchun 1 ta "boshqa" bank hisoblanadi.
    assert captured["other_bank_count"] == 1
    # language yuborilmagan bo'lsa, ExplainProductRequest standart "uz"ni
    # ishlatadi.
    assert captured["language"] == "uz"


def test_explain_product_passes_through_the_requested_language(client, monkeypatch):
    captured = {}

    def fake_explain(category, product, other_bank_count, language="uz"):
        captured["language"] = language
        return "test explanation"

    monkeypatch.setattr(api_main, "explain_featured_product", fake_explain)

    response = client.post("/explain-product", json={
        "category": "mikroqarz",
        "bank": "HamkorBank",
        "product_name": "Hamkor Mikroqarz",
        "rate_min": 10.0,
        "rate_max": 15.0,
        "term_min_months": 12,
        "term_max_months": 36,
        "amount_max_som": 100_000_000,
        "requires_collateral": False,
        "down_payment_pct": None,
        "language": "ru",
    })

    assert response.status_code == 200
    assert captured["language"] == "ru"


def test_list_categories_returns_eleven_entries(client):
    response = client.get("/categories")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 11
    keys = {c["key"] for c in data}
    assert "avtokredit" in keys
    assert "ipoteka_davlat" in keys
    assert data[0]["schema"] == "credit_down_payment"


def test_products_response_includes_new_optional_fields(client):
    response = client.get("/products", params={"category": "mikroqarz"})
    assert response.status_code == 200
    data = response.json()
    assert data[0]["down_payment_pct"] is None
    assert data[0]["grace_period_months"] is None
    assert data[0]["payment_method"] is None
    assert data[0]["special_terms"] is None


def test_list_unavailable_banks_returns_tbc_for_avtokredit(client):
    response = client.get("/unavailable-banks", params={"category": "avtokredit"})
    assert response.status_code == 200
    data = response.json()
    assert data == [{"bank": "TBC Bank", "reason": "Mahsulot mavjud emas"}]


def test_list_unavailable_banks_returns_sqb_for_avtokredit_ikkilamchi(client):
    response = client.get("/unavailable-banks", params={"category": "avtokredit_ikkilamchi"})
    assert response.status_code == 200
    data = response.json()
    assert data == [{"bank": "SQB", "reason": "Vaqtincha to'xtatilgan"}]


def test_list_unavailable_banks_returns_kapitalbank_for_avtokredit_brend_birlamchi(client):
    response = client.get("/unavailable-banks", params={"category": "avtokredit_brend_birlamchi"})
    assert response.status_code == 200
    data = response.json()
    assert data == [
        {"bank": "Kapitalbank", "reason": "Vaqtincha to'xtatilgan"},
        {"bank": "SQB", "reason": "Vaqtincha to'xtatilgan"},
    ]


def test_list_unavailable_banks_returns_empty_for_unlisted_category(client):
    response = client.get("/unavailable-banks", params={"category": "mikroqarz"})
    assert response.status_code == 200
    assert response.json() == []


def test_cors_allows_configured_frontend_origin(client):
    response = client.get(
        "/products",
        params={"category": "mikroqarz"},
        headers={"Origin": "http://localhost:5173"},
    )
    assert response.headers["access-control-allow-origin"] == "http://localhost:5173"


def test_export_excel_returns_a_valid_workbook_for_the_seeded_category(client):
    response = client.get("/export-excel", params={"category": "mikroqarz"})

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == 'attachment; filename="mikroqarz.xlsx"'

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header_row == ("#", "Bank", "Mahsulot", "Stavka", "Muddat", "Kredit miqdori", "To'lov usuli")
    data_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row == (1, "SQB", "SQB Mikroqarz", "28.0%–31.0%", "3–36 oy", "100.0 mln so'm", "Annuitet, Differensial")


def test_export_excel_translates_headers_and_values_when_language_is_ru(client):
    response = client.get("/export-excel", params={"category": "mikroqarz", "language": "ru"})

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header_row == ("#", "Банк", "Продукт", "Ставка", "Срок", "Сумма кредита", "Способ оплаты")
    data_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row[-1] == "Аннуитет, Дифференцированный"
    assert data_row[4] == "3–36 мес."


def test_export_excel_returns_404_for_an_unknown_category(client):
    response = client.get("/export-excel", params={"category": "not_a_real_category"})
    assert response.status_code == 404


def test_export_excel_all_returns_one_sheet_per_category(client):
    from categories import CATEGORIES

    response = client.get("/export-excel-all")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        'attachment; filename="bozor-tahlili-barcha-kategoriyalar.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content))
    assert len(workbook.sheetnames) == len(CATEGORIES)

    # Fixture'da faqat "mikroqarz" uchun SQB qatori bor — o'sha varaqda
    # haqiqiy ma'lumot, qolganlari faqat sarlavha qatori bilan bo'sh.
    mikroqarz_sheet = workbook[CATEGORIES[[c.key for c in CATEGORIES].index("mikroqarz")].label_uz[:31]]
    data_row = next(mikroqarz_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row[1] == "SQB"


def test_export_excel_all_translates_headers_when_language_is_ru(client):
    from categories import CATEGORIES

    response = client.get("/export-excel-all", params={"language": "ru"})

    workbook = load_workbook(BytesIO(response.content))
    mikroqarz_sheet = workbook[CATEGORIES[[c.key for c in CATEGORIES].index("mikroqarz")].label_uz[:31]]
    header_row = next(mikroqarz_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header_row == ("#", "Банк", "Продукт", "Ставка", "Срок", "Сумма кредита", "Способ оплаты")
