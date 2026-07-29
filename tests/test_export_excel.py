from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from db.models import ProductRow
from export_excel import build_all_categories_workbook, build_category_workbook
from unavailable_products import UnavailableBank


def make_product(**overrides) -> ProductRow:
    defaults = dict(
        bank="HamkorBank",
        category="avtokredit",
        product_name="Auto DAMAS",
        rate_min=0.0,
        rate_max=19.0,
        term_min_months=13,
        term_max_months=60,
        amount_max_som=600_000_000,
        requires_collateral=True,
        down_payment_pct=25.0,
        source_url="https://hamkorbank.uz",
        scraped_at=datetime.now(timezone.utc),
        grace_period_months=None,
        payment_method=None,
        special_terms=None,
    )
    defaults.update(overrides)
    return ProductRow(**defaults)


def test_build_category_workbook_sorts_by_rate_min_and_uses_down_payment_schema():
    products = [
        make_product(bank="NBU", rate_min=22.0, rate_max=22.0),
        make_product(bank="Ipoteka Bank", rate_min=0.0, rate_max=18.0),
    ]

    content = build_category_workbook(
        category_key="avtokredit",
        sheet_title="Avtokredit (birlamchi bozor)",
        products=products,
        unavailable_banks=[],
        schema="credit_down_payment",
    )

    sheet = load_workbook(BytesIO(content)).active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header == (
        "#", "Bank", "Mahsulot", "Stavka", "Muddat", "Boshlang'ich badal",
        "Imtiyozli davr", "Kredit miqdori", "To'lov usuli",
    )
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # Ipoteka Bank (0%) eng past stavka bo'lgani uchun birinchi qatorda.
    assert rows[0][:3] == (1, "Ipoteka Bank", "Auto DAMAS")
    assert rows[1][:3] == (2, "NBU", "Auto DAMAS")
    # payment_method=None -> standart "Annuitet, Differensial" ko'rsatiladi.
    assert rows[0][-1] == "Annuitet, Differensial"
    # grace_period_months=None -> "Yo'q" (noma'lum ham "yo'q" deb ko'rsatiladi).
    assert rows[0][6] == "Yo'q"


def test_build_category_workbook_uses_special_terms_schema_for_credit_special_terms():
    products = [make_product(category="istemol_krediti", special_terms="Kredit yuklamasi hisobga olinadi")]

    content = build_category_workbook(
        category_key="istemol_krediti",
        sheet_title="Iste'mol krediti",
        products=products,
        unavailable_banks=[],
        schema="credit_special_terms",
    )

    sheet = load_workbook(BytesIO(content)).active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header == (
        "#", "Bank", "Mahsulot", "Stavka", "Muddat", "Kredit miqdori",
        "Imtiyozli davr", "Maxsus shartlari", "To'lov usuli",
    )
    row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[-2] == "Kredit yuklamasi hisobga olinadi"


def test_build_category_workbook_omits_grace_period_and_special_terms_for_mikroqarz():
    products = [make_product(category="mikroqarz", product_name="Mikrokredit Plus")]

    content = build_category_workbook(
        category_key="mikroqarz",
        sheet_title="Mikroqarz (oflayn)",
        products=products,
        unavailable_banks=[],
        schema="credit_special_terms",
    )

    sheet = load_workbook(BytesIO(content)).active
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header == ("#", "Bank", "Mahsulot", "Stavka", "Muddat", "Kredit miqdori", "To'lov usuli")


def test_build_category_workbook_appends_unavailable_banks_as_extra_rows():
    content = build_category_workbook(
        category_key="avtokredit",
        sheet_title="Avtokredit (birlamchi bozor)",
        products=[make_product()],
        unavailable_banks=[UnavailableBank(bank="TBC Bank", reason="Mahsulot mavjud emas")],
        schema="credit_down_payment",
    )

    sheet = load_workbook(BytesIO(content)).active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[-1][1] == "TBC Bank"
    assert rows[-1][2] == "Mahsulot mavjud emas"


def test_build_category_workbook_sheet_title_is_truncated_to_excels_31_char_limit():
    long_title = "Ipoteka krediti (Iqtisodiyot va moliya vazirligi mablag'lari hisobidan)"
    content = build_category_workbook(
        category_key="ipoteka_davlat",
        sheet_title=long_title,
        products=[make_product()],
        unavailable_banks=[],
        schema="credit_down_payment",
    )

    sheet = load_workbook(BytesIO(content)).active
    assert len(sheet.title) <= 31


def test_build_all_categories_workbook_creates_one_named_sheet_per_category():
    categories = [
        ("avtokredit", "Avtokredit (birlamchi bozor)", "credit_down_payment"),
        ("mikroqarz", "Mikroqarz (oflayn)", "credit_special_terms"),
    ]
    products_by_category = {
        "avtokredit": [make_product(bank="HamkorBank")],
        "mikroqarz": [make_product(category="mikroqarz", product_name="Mikrokredit Plus")],
    }

    content = build_all_categories_workbook(
        categories=categories,
        products_by_category=products_by_category,
        unavailable_by_category={},
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["Avtokredit (birlamchi bozor)", "Mikroqarz (oflayn)"]

    avto_row = next(workbook["Avtokredit (birlamchi bozor)"].iter_rows(min_row=2, max_row=2, values_only=True))
    assert avto_row[1] == "HamkorBank"

    mikro_header = next(workbook["Mikroqarz (oflayn)"].iter_rows(min_row=1, max_row=1, values_only=True))
    assert mikro_header == ("#", "Bank", "Mahsulot", "Stavka", "Muddat", "Kredit miqdori", "To'lov usuli")


def test_build_all_categories_workbook_handles_a_category_with_no_products():
    categories = [("avtokredit", "Avtokredit (birlamchi bozor)", "credit_down_payment")]

    content = build_all_categories_workbook(
        categories=categories,
        products_by_category={},
        unavailable_by_category={"avtokredit": [UnavailableBank(bank="TBC Bank", reason="Mahsulot mavjud emas")]},
    )

    sheet = load_workbook(BytesIO(content)).active
    row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert row[1] == "TBC Bank"


def test_build_all_categories_workbook_deduplicates_sheet_titles_that_collide_after_truncation():
    long_prefix = "A" * 40
    categories = [
        ("cat_one", f"{long_prefix} one", "credit_down_payment"),
        ("cat_two", f"{long_prefix} two", "credit_down_payment"),
    ]

    content = build_all_categories_workbook(
        categories=categories,
        products_by_category={"cat_one": [make_product()], "cat_two": [make_product()]},
        unavailable_by_category={},
    )

    sheet_names = load_workbook(BytesIO(content)).sheetnames
    assert len(sheet_names) == len(set(sheet_names)) == 2
